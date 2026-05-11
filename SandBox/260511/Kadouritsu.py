import time
import os
import sys
import pandas as pd
import warnings
import logging
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime

# 警告・ログの抑制
warnings.filterwarnings('ignore')
logging.getLogger('xlrd').setLevel(logging.ERROR)

# --- 設定項目 ---
LOGIN_URL = "http://itskikan/"
DL_PAGE_URL_TEMPLATE = "http://10.76.94.45/kosucheck/kadoritu_mei.php?nengetu={year_month}&bumoncd=100014&bumonknj=%E5%95%86%E5%93%81C%201%E9%83%A8"

USER_ID = "7004141"
PASSWORD = "Hrykszk=01"

DOWNLOAD_DIR = os.path.join(os.getcwd(), "temp_download")
TARGET_EXCEL = r"C:\Users\7004141\OneDrive - Panasonic\作業用\稼働率.xlsx"
TARGET_SHEET = datetime.now().strftime("%Y%m")

CHROME_EXE = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
#DRIVER_EXE = r"c:\Users\7004141\OneDrive - Panasonic\作業用\Git-Work\suzu76\SandBox\260511\chromedriver.exe"

def automate_full_task():
    # 0. Excelプロセスを強制終了（書き込みエラー防止）
    #os.system(f'taskkill /F /IM EXCEL.EXE /T > NUL 2>&1')

    # Z. 開始時間を記録
    start_time = time.time()

    if not os.path.exists(DOWNLOAD_DIR): 
        os.makedirs(DOWNLOAD_DIR)
    
    # --- ブラウザ設定 ---
    options = Options()
    options.binary_location = CHROME_EXE
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--unsafely-treat-insecure-origin-as-secure=http://10.76.94.45')
    
    # ダウンロード設定の集約
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "safebrowsing.disable_download_protection": True
    }
    options.add_experimental_option("prefs", prefs)

    #service = Service(executable_path=DRIVER_EXE)
    #driver = webdriver.Chrome(service=service, options=options)
    # これだけで Selenium Manager が自動起動します
    service = Service() 
    driver = webdriver.Chrome(service=service, options=options)
    
    # Headlessモードでのダウンロードを明示的に許可
    driver.execute_cdp_cmd("Page.setDownloadBehavior", {
        "behavior": "allow", "downloadPath": DOWNLOAD_DIR
    })
    
    try:
        wait = WebDriverWait(driver, 15)

        # 1. ログイン処理
        print("🔑 ログイン中...")
        driver.get(LOGIN_URL)
        
        uid_field = wait.until(EC.presence_of_element_located((By.NAME, "uid")))
        uid_field.send_keys(USER_ID)
        driver.find_element(By.NAME, "pass").send_keys(PASSWORD)
        driver.find_element(By.NAME, "loginbtn").click()
        
        # ログイン後の安定を待つ
        time.sleep(2) 
        print("✅ ログイン完了")

        # 2. 第1階層：稼動率リンクをクリック
        print("🖱️ 「稼動率（係長以上）」へ移動中...")
        target_link = wait.until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "稼動率（係長以上）")))
        driver.execute_script("arguments[0].scrollIntoView(true);", target_link)
        time.sleep(1)
        target_link.click()
        
        # 3. 第2階層：部門別明細リンクをクリック
        print("🖱️ PDC1部の「明細」へ移動中...")
        target_xpath = '//a[contains(@href, "bumoncd=100014") and text()="明細"]'
        meisai_link = wait.until(EC.presence_of_element_located((By.XPATH, target_xpath)))
        driver.execute_script("arguments[0].click();", meisai_link)
        
        # 4. 第3階層：EXCELダウンロード
        try:
            target_xpath = '//input[@type="button" and @value="EXCEL"] | //input[@type="submit" and @value="EXCEL"]'
            excel_btn = wait.until(EC.presence_of_element_located((By.XPATH, target_xpath))) 
            driver.execute_script("arguments[0].click();", excel_btn)          
            print("⏳ ダウンロード中...")

        # ★追加：ダウンロードが完了するまで最大30秒待機する
            timeout = 30
            for i in range(timeout):
                # フォルダ内に .xlsx または .xls が出現し、かつ .crdownload（一時ファイル）が無くなるまで待つ
                files = os.listdir(DOWNLOAD_DIR)
                downloading = any(f.endswith('.crdownload') for f in files)
                excel_exists = any(f.endswith(('.xlsx', '.xls')) for f in files)
                
                if excel_exists and not downloading:
                    break
                time.sleep(1)

        except Exception as e:
            print("❌ EXCELボタン操作でエラーが発生")
            raise e

        # 5. 最新ファイルの取得
        files = [f for f in os.listdir(DOWNLOAD_DIR) if f.endswith(('.xlsx', '.xls'))]
        if not files:
            print("❌ ファイルが見つかりませんでした。")
            return
        
        latest_file = os.path.join(DOWNLOAD_DIR, sorted(files)[-1])
        print(f"✨ ダウンロード完了！: {os.path.basename(latest_file)}")

        # 6. データ抽出（警告消しゴム機能付き）
        print("📊 データ抽出中...", end="\r")
        df = pd.read_excel(latest_file, usecols="A:K", nrows=49, skiprows=2, header=None)
        #with open(latest_file, 'rb') as f:
        #    df = pd.read_excel(f, usecols="A:K", nrows=49, skiprows=2, header=None)
        
        # 警告行を上書き消去
        sys.stdout.write("\033[F\033[K") 
        sys.stdout.flush()

        # 7. 値のみ転記
        print("📝 転記中...")
        book = load_workbook(TARGET_EXCEL)
        if TARGET_SHEET in book.sheetnames:
            sheet = book[TARGET_SHEET]
            
            # DataFrameの行数と列数を取得
            rows, cols = df.shape
            
            for r in range(rows):
                for c in range(cols):
                    # 書き込む値を取得（NaNなどは空文字に変換するとより安全）
                    val = df.iloc[r, c]
                    if pd.isna(val):
                        val = None
                    
                    # セルを特定（startrow=3, startcol=1 に合わせる）
                    target_cell = sheet.cell(row=r + 3, column=c + 1)
                    
                    # ★重要：ここが「値だけ」をセットする命令です
                    target_cell.value = val
            
            # 保存時に書式情報が壊れないよう明示的に保存
            book.save(TARGET_EXCEL)
            print("✨ 転記完了！")
        else:
            print(f"❌ シート '{TARGET_SHEET}' が見つかりません。")

        # 8. ファイル起動
        print(f"📂 更新したファイルオープン: {os.path.basename(TARGET_EXCEL)}")
        os.startfile(TARGET_EXCEL)

        # Z. 終了時間を記録（正常終了の直前）
        elapsed_sec = time.time() - start_time
        print(f"\n⏱️ 実行時間: {elapsed_sec:.2f} 秒")

    except Exception as e:
        print(f"🔥 エラー発生: {e}")
    
    finally:
        driver.quit()
        time.sleep(1) # 解放を待つための「一呼吸」

        # お掃除実行
        if os.path.exists(DOWNLOAD_DIR):
            for f in os.listdir(DOWNLOAD_DIR):
                file_path = os.path.join(DOWNLOAD_DIR, f)
                try:
                    if os.path.isfile(file_path):
                        os.remove(file_path)
                except Exception as e:
                    # まだ消せない場合はさらに1秒待ってリトライ
                    time.sleep(1)
                    try:
                        os.remove(file_path)
                    except:
                        print(f"⚠️ ファイル {f} の削除をスキップしました（使用中）")

if __name__ == "__main__":
    automate_full_task()